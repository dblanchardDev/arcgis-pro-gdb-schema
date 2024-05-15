"""
Utility to quickly add a mandatory dropdown to a cell in a consistent fashion.

Author: David Blanchard
"""

import re
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell


class CellDropdown:
    """Prepare a re-usable cell dropdown configuration which will add an enforced dropdown on cells.

    Args:
        workbook (openpyxl.Workbook): Workbook in which the dropdown will be added.
        values (list[str]): List of values to be included in dropdown.
        label (str): Label used when displaying an error message to users.
        user_fill_in (bool, optional): Allow users to fill-in their own values, other than what the dropdown contains. Defaults to False.
        message (str, optional): The message displayed when a value other than one in the dropdown is selected. Defaults to "Choose a {label} from the dropdown".
    """ #pylint: disable=line-too-long

    _label:str = None
    _user_fill_in:bool = False
    _message:bool = False
    _range_name:str = None


    def __init__(self, workbook:openpyxl.Workbook, values:list[str], label:str, user_fill_in:bool=False,
                 message:str=None) -> None:
        self._label = label
        self._user_fill_in = user_fill_in
        self._message = message

        sheet = self._get_lookup_sheet(workbook)
        column_number = self._find_next_empty_column(sheet)

        # Add the label followed by each value
        sheet.cell(1, column_number).value = label

        for idx, val in enumerate(values):
            sheet.cell(idx + 2, column_number).value = val

        # Create a named range
        self._range_name = re.sub(r"[^\w\d_]", "_", label)
        self._range_name = re.sub(r"_+", "_", self._range_name)
        self._range_name = "dropdowns." + self._range_name.lower()

        column_letter = openpyxl.utils.get_column_letter(column_number)
        last_row = len(values) + 1
        workbook.create_named_range(self._range_name, sheet, f"${column_letter}$2:${column_letter}${last_row}")


    def _get_lookup_sheet(self, workbook:openpyxl.Workbook) -> Worksheet:
        """Get a reference to the dropdown lookup sheet. If it does not exit yet, create it.

        Args:
            workbook (openpyxl.Workbook): The workbook that (will) contains the dropdown lookup sheet.

        Returns:
            Worksheet: Reference to the worksheet.
        """
        sheet_name = "_dropdown_lookup_"

        if sheet_name not in workbook:
            sheet:Worksheet = workbook.create_sheet(sheet_name, 0)
            sheet.sheet_state = Worksheet.SHEETSTATE_HIDDEN

        return workbook[sheet_name]


    def _find_next_empty_column(self, sheet:Worksheet) -> int:
        """Find the next column that has no text in the top row.

        Args:
            sheet (Worksheet): Worksheet in which the column is to be located.

        Raises:
            NotImplementedError: Exceeded 50 columns.

        Returns:
            int: Number of the column, left to right, starting at 1.
        """
        column_number:int = None
        cell:"Cell" = None
        for index in range(50):
            cell = sheet.cell(1, index + 1)
            if cell.value is None:
                column_number = index + 1
                break
        else:
            raise NotImplementedError("Adding more than 50 different dropdowns isn't supported.")

        return column_number


    def data_validation(self) -> DataValidation:
        """Get a openpyxl Data Validation object for this dropdown.

        Returns:
            DataValidation: Data validation object applying a dropdown.
        """

        message = f"Choose a valid {self._label.lower()} from the dropdown."
        error_title = f"Invalid {self._label.title()}"
        error_style = "stop"

        if self._user_fill_in:
            message = f"A user-defined value has been entered for {self._label.lower()}."
            error_title = f"User-defined {self._label.title()}"
            error_style = "information"

        if self._message:
            message = self._message


        return DataValidation(
            type="list",
            formula1=f"={self._range_name}",
            showInputMessage=False,
            allow_blank=True,
            errorTitle=error_title,
            error=message,
            errorStyle=error_style,
        )


_dropdowns_yes_no = {}

def get_dropdown_boolean(workbook:openpyxl.Workbook) -> CellDropdown:
    """Get a yes/no dropdown which can be shared between different sections of a workbook.

    Args:
        workbook (openpyxl.Workbook): The workbook to which the dropdown will belong.

    Returns:
        CellDropdown: The dropdown cell modifier.
    """

    if workbook not in _dropdowns_yes_no:
        _dropdowns_yes_no[workbook] = CellDropdown(
            workbook=workbook,
            values=[
                "yes",
                "no",
            ],
            label="boolean",
        )

    return _dropdowns_yes_no[workbook]
