"""
Utility class to write to the cells of a specific worksheet.

Author: David Blanchard
"""

import datetime
from typing import TYPE_CHECKING, Union

import openpyxl

from .cell_dropdown import CellDropdown
from .cell_options import CELL_OPTS_DEFAULT, CellOptions
from .cell_rule import CellRule

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# Custom Types
CellValues = Union[str, float, int, datetime.datetime, "MERGE_CELL_WITH_LEFT"]
CellDescriptor = Union[CellValues, tuple[CellValues, Union[CellOptions, CellRule]]]
ValueBlock = list[list[CellDescriptor]]
DefaultCellOptions = Union[CellOptions, list[CellOptions]]


#pylint: disable=too-few-public-methods
class _MergeCellWithLeft:
    """Flag a cell as being merged with the one to the left of the current cell."""

MERGE_CELL_WITH_LEFT = _MergeCellWithLeft()


class CellWriter:
    """Utility to write to the cells of a specific worksheet.

    Args:
        worksheet (Worksheet): Worksheet to which values will be written.
    """

    _worksheet:"Worksheet" = None
    _next_row:int = 1
    _next_col:int = 1
    _table_style_saved:openpyxl.worksheet.table.TableStyleInfo = None


    def __init__(self, worksheet:"Worksheet") -> None:
        self._worksheet = worksheet


    @property
    def worksheet(self) -> "Worksheet":
        """Worksheet in which the cell writing will occur."""
        return self._worksheet


    def write_cell(self, value:CellValues, options:CellOptions=None, rules:Union[CellRule, list[CellRule]]=None,
                   dropdown:CellDropdown=None, row:int=None, col:int=1) -> tuple[int]:
        """Write a value to particular cell.

        Args:
            value (CellValues): Value to be written to cell.
            options (CellOptions, optional): Optional settings for the cell. Defaults to None.
            rules (Union[CellRule, list[CellRule]], optional): Cell rules that apply to this cell. Provide 1 rule or a list of rules. Defaults to None.
            dropdown (CellDropdown, optional): A dropdown to apply to this cell. Defaults to None.
            row (int, optional): Row number starting on top with 1. Defaults to the next row that follows the previously edited cell.
            col (str, optional): Column number starting on left with 1. Defaults to the first column. To get the next column that follows the previously edited cell, set to None

        Returns:
            tuple[int]: Coordinates of the cell that was edited (row, column).
        """ #pylint: disable=line-too-long

        # Set defaults if none specified in parameters
        if value is None:
            value = ""
        if options is None:
            options = CELL_OPTS_DEFAULT
        if row is None:
            row = self._next_row
        if col is None:
            col = self._next_col

        # Get reference to cell
        cell = self.worksheet.cell(row, col)

        # Set the value and other key properties
        cell.value = value
        cell.style = options.style
        cell.number_format = options.num_format

        if options.link:
            cell.hyperlink = options.link

        # Update the row and column tracking
        self._next_row = row + 1
        self._next_col = col + 1

        # Apply Rules
        if rules:
            rule_list = rules if rules is isinstance(rules, (list, tuple)) else rules
            for rule in rule_list:
                self.add_rule(rule, row, col)

        # Apply dropdown
        if dropdown:
            self.add_dropdown(dropdown, row, col)

        # Return the row and column coordinates
        return (row, col)


    def _write_value_block_cell(self, cell_data:CellValues, row:int, col:int, default_options:CellOptions=None):

        # Complete a merge with left
        if isinstance(cell_data, _MergeCellWithLeft):
            self.worksheet.merge_cells(
                start_row=row,
                start_column=col - 1,
                end_row=row,
                end_column=col,
            )

        # Deal a sequence that includes cell options/rules/dropdowns
        elif isinstance(cell_data, (tuple, list)):
            value = cell_data[0]
            options = default_options
            rules = []
            dropdown = None

            for attribute in cell_data[1:]:
                if isinstance(attribute, CellOptions):
                    options = attribute
                elif isinstance(attribute, CellRule):
                    rules.append(attribute)
                elif isinstance(attribute, CellDropdown):
                    dropdown = attribute

            self.write_cell(value, options, rules, dropdown, row, col)

        # Deal with singleton values
        else:
            self.write_cell(
                value=cell_data,
                options=default_options,
                row=row,
                col=col,
            )


    def write_block(self, value_block:ValueBlock, start_row:int=None, start_col:int=1,
                    default_cell_options:DefaultCellOptions=None) -> tuple[tuple[int]]:
        """Write several rows and columns at the same time by passing in a matrix of values.

        Args:
            value_block (ValueBlock): A list of rows, itself containing a list of columns, itself containing a cell value (e.g. [["A1", "B1", "C3"], ["A2", "B2"]]). Instead of the cell value, you may specify a cell descriptor, which is a tuple whose first value is the cell value, followed by up to 1 CellOptions, up to 1 CellDropdown, and as many CellRule as required. Order of CellOptions, CellRule, and CellDropdown does not matter. You may also pass MERGE_CELL_WITH_LEFT to merge this cell.
            start_row (int, optional): The top-most row number where the value block starts. Defaults to the row that follows the previously edited cell.
            start_col (int, optional): The left-most column number where the value block starts. Defaults to the left-most column. Set to None to populate the column that follows the previously edited one.
            default_cell_options (DefaultCellOptions): Cell options that will be applied to all cells when the value_block doesn't contain one. To specify different options by column, use a list ordered by column from left to right.

        Returns:
            tuple[tuple[int]]: Range edited from top-left to bottom-right ((start row, start col), (end row, end col))
        """ #pylint: disable=line-too-long

        # Set defaults if none specified in parameters
        if start_row is None:
            start_row = self._next_row
        if start_col is None:
            start_col = self._next_col

        # Loop through the data
        end_row = row = start_row
        end_col = col = start_col

        for row_data in value_block:
            for cell_data in row_data:

                # Pick-out the default cell option for this column
                default_options = None
                if isinstance(default_cell_options, CellOptions):
                    default_options = default_cell_options
                elif isinstance(default_cell_options, (list, tuple)) and len(default_cell_options) >= col:
                    default_options = default_cell_options[col - 1]

                # Write out the row
                self._write_value_block_cell(cell_data, row, col, default_options)

                # Update the end values when necessary
                end_row = max(end_row, row)
                end_col = max(end_col, col)

                # Increase column count after each cell
                col += 1

            # Increase row count and reset column after each row
            row += 1
            col = start_col

        return ((start_row, start_col), (end_row, end_col))


    def set_widths(self, widths:list[int]):
        """Set the width of all columns.

        Args:
            widths (list[int]): List of widths starting from the leftmost column.
        """
        for idx, w in enumerate(widths):
            col_letter = openpyxl.utils.get_column_letter(idx + 1)
            self.worksheet.column_dimensions[col_letter].width = w


    def increase_basic_height(self, row:int, factor:int):
        """Set the row height to be a multiple of the height of a basic style cell.

        Args:
            row (int): Number of the row, starting at 1 at the top.
            factor (int): Number of times the height it to be increased by.
        """
        self._worksheet.row_dimensions[row].height = 13.2 * factor


    @property
    def _table_style(self):
        if self._table_style_saved is None:
            self._table_style_saved = openpyxl.worksheet.table.TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

        return self._table_style_saved


    def make_table(self, top_row:int, left_col:int, bot_row:int, right_col:int,
                   name:str="") -> openpyxl.worksheet.table.Table:
        """Make a range a formatted table in Excel.

        Args:
            top_row (int): Top-most row number.
            left_col (int): Left-most column number.
            bot_row (int): Bottom-most row number.
            right_col (int): Right-most column number.
            name (str, optional): Name given to the table. Defaults to "".
        Returns:
            openpyxl.worksheet.table.Table : The table that was created.
        """
        left_col_letter = openpyxl.utils.get_column_letter(left_col)
        right_col_letter = openpyxl.utils.get_column_letter(right_col)
        ref = f"{left_col_letter}{top_row}:{right_col_letter}{bot_row}"

        table = openpyxl.worksheet.table.Table(ref=ref, displayName=name, tableStyleInfo=self._table_style)
        self.worksheet.add_table(table)
        return table


    def add_rule(self, rule:CellRule, top_row:int, left_col:int, bot_row:int=None, right_col:int=None):
        """Add a cell value conditional formatting rule to a range of cells.

        Args:
            rule (CellRule): Cell Rule to apply.
            top_row (int): Top-most row number.
            left_col (int): Left-most column number.
            bot_row (int, optional): Bottom-most row number. Defaults to None.
            right_col (int, optional): Right-most column number. Defaults to None.
        """

        left_col_letter = openpyxl.utils.get_column_letter(left_col)
        range_string = f"{left_col_letter}{top_row}"

        if left_col and bot_row:
            right_col_letter = openpyxl.utils.get_column_letter(right_col)
            range_string += f":{right_col_letter}{bot_row}"

        base_coordinate = f"{left_col_letter}{top_row}"
        self.worksheet.conditional_formatting.add(range_string, rule.make_rule_for_cell(base_coordinate))


    def add_dropdown(self, dropdown:CellDropdown, top_row:int, left_col:int, bot_row:int=None, right_col:int=None):
        """Add a cell dropdown data validation to a range of cells.

        Args:
            dropdown (CellDropdown): Cell Dropdown to apply.
            top_row (int): Top-most row number.
            left_col (int): Left-most column number.
            bot_row (int, optional): Bottom-most row number. Defaults to None.
            right_col (int, optional): Right-most column number. Defaults to None.
        """

        left_col_letter = openpyxl.utils.get_column_letter(left_col)
        range_string = f"{left_col_letter}{top_row}"

        if left_col and bot_row:
            right_col_letter = openpyxl.utils.get_column_letter(right_col)
            range_string += f":{right_col_letter}{bot_row}"

        dv = dropdown.data_validation()
        self.worksheet.add_data_validation(dv)
        dv.add(range_string)