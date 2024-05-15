"""
Parse the values found in the Excel workbook into a standard set of dictionaries.

Author: Roya Shourouni
"""

from typing import TypedDict, Union

import openpyxl


# Types
CellValues = Union[str, int, float, None]
KeyValuePair = dict[str, CellValues]

class SheetRoot(TypedDict):
    """Type for Excel to data results."""
    key_value_pairs: KeyValuePair
    tables: dict[str, list[KeyValuePair]]

WorksheetList = list[SheetRoot]


def parse_workbook(workbook_path:str=None) -> tuple[KeyValuePair, WorksheetList]:
    """Parse an entire schema workbook into a simple data structure.

    Args:
        workbook_path (str, optional): Path to the workbook to be read. Will return empty data if not specified.
        Defaults to None.

    Returns:
        tuple[KeyValuePair, WorksheetList]: A tuple containing the data as it was read (gdb info sheet, sheet data)
    """

    # 1. Open the workbook -- return empty data if no workbook provided.
    wb = openpyxl.load_workbook(workbook_path)
    gdb_info = {}
    sheet_data = []

    for sheetname in wb.sheetnames:
        if sheetname == '_INFO_' or (not sheetname.startswith("_") and not sheetname.endswith("_")):
            sheet = wb[sheetname]
            sheet_dict, key_value_pairs, tables, table_list, column_dict = {},{}, {}, [], {}
            try:
                for table in sheet.tables.values():
                    table_range = table.ref  # retrieve the range of table
                    heading = sheet[table_range][0][0]
                    table_name = sheet.cell(row=heading.row -1 , column=heading.column).value

                    # 2. Read each sheet's tables and create a table_list
                    for row in sheet[table_range][1:]: # skip the first row of table which is header
                        for column in table.tableColumns: # retrieve the header/column names
                            col_index = table.tableColumns.index(column)
                            column_dict[column.name]=row[col_index].value # assign the cell of row to each header

                        #Skip empty rows
                        row_is_empty = True
                        for k,v in column_dict.items():
                            if k != 'Order' and v is not None:
                                row_is_empty = False
                                break

                        if not row_is_empty:
                            table_list.append(column_dict)

                        column_dict = {}
                    tables[table_name] = table_list
                    table_list = []


                # Create a dict of each sheet's key-value pairs from all rows that have a value in column A
                # and are not contained within a table
                    for row in sheet[table_range]:
                        for cell in row:
                            cell.value = None  # Clear cells that are within a table
                for key, *values in sheet.iter_rows(min_col=1, max_col=2):
                    if key.value is not None:
                        key_value_pairs[key.value] = values[0].value

                # 3. Read the GDB info from the _INFO_ sheet's key-value pairs
                if sheetname == '_INFO_':
                    gdb_info = key_value_pairs
                else:
                    sheet_dict["key_value_pairs"] = key_value_pairs
                    sheet_dict["tables"] = tables
                    sheet_data.append(sheet_dict)
            except Exception as exc:
                #pylint: disable-next=broad-exception-raised
                raise Exception(f"Failure while reading {sheetname} spreadsheet.") from exc

    return (gdb_info, sheet_data)